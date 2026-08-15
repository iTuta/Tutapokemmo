from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "web" / "search-data.js"
OUTPUT = ROOT / "PokeMMO_进化链与稀有地点分析.xlsx"

F = {
    "ID": 0,
    "NAME": 1,
    "BASE": 2,
    "TYPES": 3,
    "REGION": 4,
    "LOC": 5,
    "TERRAIN": 6,
    "LEVEL": 7,
    "SEASON": 8,
    "HORDE": 9,
    "MORNING": 10,
    "DAY": 11,
    "NIGHT": 12,
    "R_MORNING": 13,
    "R_DAY": 14,
    "R_NIGHT": 15,
    "FORM": 16,
}

TIME_FIELDS = [
    ("早晨", F["MORNING"], F["R_MORNING"]),
    ("白天", F["DAY"], F["R_DAY"]),
    ("夜晚", F["NIGHT"], F["R_NIGHT"]),
]
REGION_ORDER = {name: index for index, name in enumerate(["关都", "丰缘", "神奥", "合众", "城都"])}
TIER_ORDER = {f"T{i}": i for i in range(8)}
TIER_SCORES = {"T0": 50, "T1": 45, "T2": 40, "T3": 30, "T4": 15, "T5": 10, "T6": 5, "T7": 3}
RARE_TIERS = {"T0", "T1", "T2", "T3"}
MISSING_TIER_SEEDS = {
    "T3": [331, 434, 441, 443, 453, 456, 554, 556, 561, 615],
    "T6": [304, 459, 504, 517, 522, 524, 527, 535, 540, 572, 574, 577, 582, 599, 605, 618, 621, 629, 632],
    "T7": [422, 431, 436, 550, 551, 562, 585, 592, 607, 619, 622],
}

COLORS = {
    "navy": "1F4E78",
    "blue": "D9EAF7",
    "teal": "0F6B78",
    "green": "DDEBF7",
    "gold": "F4B183",
    "night": "E4DFEC",
    "light": "F7F9FB",
    "border": "C7D0D9",
    "white": "FFFFFF",
    "text": "1F2937",
}
TIER_COLORS = {
    "T0": "FFD966",
    "T1": "D9E1F2",
    "T2": "F4B183",
    "T3": "C6E0B4",
}


def load_data() -> dict:
    text = SOURCE.read_text(encoding="utf-8").strip()
    prefix = "window.POKEMON_DATA="
    if not text.startswith(prefix):
        raise ValueError(f"Unexpected data wrapper in {SOURCE}")
    payload = text[len(prefix):]
    if payload.endswith(";"):
        payload = payload[:-1]
    return json.loads(payload)


def parse_rate(value) -> float | None:
    match = re.fullmatch(r"\s*(\d+(?:\.\d+)?)%\s*", str(value or ""))
    return float(match.group(1)) if match else None


def time_info(record: list) -> tuple[str, int, bool, list[str], float | None, list[str]]:
    active_names = [name for name, active_field, _ in TIME_FIELDS if record[active_field]]
    active = tuple(bool(record[field]) for _, field, _ in TIME_FIELDS)
    if active == (False, False, True):
        label, weight = "仅夜晚", 100
    elif active[2] and not all(active):
        label, weight = "/".join(active_names), 80
    elif all(active):
        label, weight = "全天", 60
    elif active == (False, True, False):
        label, weight = "仅白天", 50
    elif active == (True, False, False):
        label, weight = "仅早晨", 40
    else:
        label, weight = "/".join(active_names) or "未标明", 30

    numeric_rates = []
    special_rates = []
    for name, active_field, rate_field in TIME_FIELDS:
        if not record[active_field]:
            continue
        rate = parse_rate(record[rate_field])
        if rate is not None:
            numeric_rates.append((name, rate))
        else:
            raw = str(record[rate_field] or "").strip()
            if raw and raw != "--":
                special_rates.append(raw)

    max_rate = max((rate for _, rate in numeric_rates), default=None)
    best_times = [name for name, rate in numeric_rates if rate == max_rate] if max_rate is not None else []
    return label, weight, active == (False, False, True), best_times, max_rate, sorted(set(special_rates))


def horde_label(value: int) -> str:
    return {0: "普通", 3: "3群怪", 5: "5群怪"}.get(value, str(value))


def candidate(record: list) -> dict:
    time_label, time_weight, night_only, best_times, max_rate, special_rates = time_info(record)
    return {
        "record": record,
        "rate": max_rate,
        "rate_text": f"{max_rate:g}%" if max_rate is not None else "/".join(special_rates) or "未标明",
        "best_times": "/".join(best_times) if best_times else "-",
        "time_label": time_label,
        "time_weight": time_weight,
        "night_only": night_only,
        "numeric": max_rate is not None,
    }


def candidate_sort_key(item: dict) -> tuple:
    record = item["record"]
    encounter_priority = 2 if record[F["HORDE"]] == 0 else 1
    return (
        1 if item["numeric"] else 0,
        item["rate"] if item["rate"] is not None else -1,
        item["time_weight"],
        encounter_priority,
        -record[F["ID"]],
    )


def choose_best(records: list[list]) -> dict:
    candidates = [candidate(record) for record in records]
    candidates.sort(
        key=lambda item: (
            candidate_sort_key(item),
            item["record"][F["REGION"]],
            item["record"][F["LOC"]],
            item["record"][F["TERRAIN"]],
            item["record"][F["SEASON"]],
        ),
        reverse=True,
    )
    return candidates[0]


def build_families(data: dict) -> tuple[dict[int, str], dict[str, list[int]], dict[str, str]]:
    id_to_family = {}
    family_ids = {}
    family_names = {}
    meta = data.get("m", {})

    for index, ids in enumerate(data.get("f", [])):
        key = f"F{index + 1:03d}"
        family_ids[key] = sorted(ids)
        for pokemon_id in ids:
            id_to_family[pokemon_id] = key

    all_ids = {record[F["ID"]] for record in data["r"]}
    for pokemon_id in sorted(all_ids):
        if pokemon_id not in id_to_family:
            key = f"S{pokemon_id:03d}"
            id_to_family[pokemon_id] = key
            family_ids[key] = [pokemon_id]

    for key, ids in family_ids.items():
        names = [meta.get(str(pokemon_id), [str(pokemon_id)])[0] for pokemon_id in ids]
        family_names[key] = " → ".join(names)
    return id_to_family, family_ids, family_names


def build_tiers(data: dict, id_to_family: dict[int, str], family_ids: dict[str, list[int]]) -> dict[int, tuple[str, int]]:
    tiers = {int(pokemon_id): (value[0], value[1]) for pokemon_id, value in data.get("ti", {}).items()}
    for tier, seeds in MISSING_TIER_SEEDS.items():
        for seed in seeds:
            family_key = id_to_family.get(seed)
            ids = family_ids.get(family_key, [seed])
            for pokemon_id in ids:
                tiers.setdefault(pokemon_id, (tier, TIER_SCORES[tier]))
    return tiers


def selected_row(
    family_key: str,
    best: dict,
    family_name: str,
    tier: tuple[str, int] | None,
    region_count: int,
    other_regions: list[str],
) -> list:
    record = best["record"]
    return [
        0,
        record[F["REGION"]],
        family_key,
        family_name,
        record[F["BASE"]],
        record[F["ID"]],
        best["rate"],
        best["rate_text"],
        best["best_times"],
        best["time_label"],
        best["time_weight"],
        "是" if best["night_only"] else "否",
        record[F["LOC"]],
        record[F["TERRAIN"]],
        record[F["LEVEL"]],
        record[F["SEASON"]],
        horde_label(record[F["HORDE"]]),
        tier[0] if tier else "-",
        tier[1] if tier else None,
        region_count,
        "、".join(other_regions) if other_regions else "-",
        "数值概率优先；同概率时仅夜晚和普通遭遇优先" if best["numeric"] else "该进化链没有数值概率，保留特殊遇见记录",
    ]


def build_selection_tables(data: dict, id_to_family: dict[int, str], family_names: dict[str, str], tiers: dict[int, tuple[str, int]]):
    records_by_family = defaultdict(list)
    records_by_family_region = defaultdict(list)
    regions_by_family = defaultdict(set)
    for record in data["r"]:
        family_key = id_to_family[record[F["ID"]]]
        records_by_family[family_key].append(record)
        records_by_family_region[(family_key, record[F["REGION"]])].append(record)
        regions_by_family[family_key].add(record[F["REGION"]])

    global_best = {key: choose_best(records) for key, records in records_by_family.items()}
    global_rows = []
    for family_key, best in global_best.items():
        record = best["record"]
        regions = sorted(regions_by_family[family_key], key=lambda name: (REGION_ORDER.get(name, 99), name))
        other_regions = [region for region in regions if region != record[F["REGION"]]]
        global_rows.append(selected_row(
            family_key,
            best,
            family_names[family_key],
            tiers.get(record[F["ID"]]),
            len(regions),
            other_regions,
        ))

    region_rows = []
    special_rows = []
    for (family_key, region), records in records_by_family_region.items():
        best = choose_best(records)
        record = best["record"]
        row = selected_row(
            family_key,
            best,
            family_names[family_key],
            tiers.get(record[F["ID"]]),
            len(regions_by_family[family_key]),
            [name for name in sorted(regions_by_family[family_key]) if name != region],
        )
        row.append("是" if global_best[family_key]["record"][F["REGION"]] == region else "否")
        region_rows.append(row)
        if not best["numeric"]:
            special_rows.append(row[:-1])

    def sort_rows(rows: list[list]) -> None:
        rows.sort(key=lambda row: (
            REGION_ORDER.get(row[1], 99),
            -row[10],
            -(row[6] if row[6] is not None else -1),
            row[3],
        ))
        rank_by_region = defaultdict(int)
        for row in rows:
            rank_by_region[row[1]] += 1
            row[0] = rank_by_region[row[1]]

    sort_rows(global_rows)
    sort_rows(region_rows)
    sort_rows(special_rows)
    return global_rows, region_rows, special_rows, global_best


def build_rare_tables(
    data: dict,
    id_to_family: dict[int, str],
    family_names: dict[str, str],
    tiers: dict[int, tuple[str, int]],
):
    grouped = defaultdict(list)
    for record in data["r"]:
        tier = tiers.get(record[F["ID"]])
        if tier and tier[0] in RARE_TIERS:
            grouped[(record[F["REGION"]], record[F["LOC"]], record[F["ID"]])].append(record)

    detail_objects = []
    for (region, location, pokemon_id), records in grouped.items():
        best = choose_best(records)
        tier, score = tiers[pokemon_id]
        active_union = [any(record[field] for record in records) for _, field, _ in TIME_FIELDS]
        night_only = active_union == [False, False, True]
        time_union = [TIME_FIELDS[i][0] for i, active in enumerate(active_union) if active]
        detail_objects.append({
            "region": region,
            "location": location,
            "pokemon_id": pokemon_id,
            "pokemon": records[0][F["BASE"]],
            "family": id_to_family[pokemon_id],
            "family_name": family_names[id_to_family[pokemon_id]],
            "tier": tier,
            "score": score,
            "rate": best["rate"],
            "rate_text": best["rate_text"],
            "best_times": best["best_times"],
            "times": "仅夜晚" if night_only else "/".join(time_union),
            "night_only": night_only,
            "seasons": "、".join(sorted({record[F["SEASON"]] for record in records})),
            "terrains": "、".join(sorted({record[F["TERRAIN"]] for record in records})),
            "hordes": "、".join(sorted({horde_label(record[F["HORDE"]]) for record in records})),
            "levels": "、".join(sorted({str(record[F["LEVEL"]]) for record in records})[:10]),
        })

    detail_objects.sort(key=lambda item: (
        REGION_ORDER.get(item["region"], 99),
        item["location"],
        TIER_ORDER[item["tier"]],
        item["pokemon"],
    ))
    detail_rows = [[
        item["region"], item["location"], item["pokemon"], item["pokemon_id"],
        item["family"], item["family_name"], item["tier"], item["score"],
        item["rate"], item["rate_text"], item["best_times"], item["times"],
        "是" if item["night_only"] else "否", item["seasons"], item["terrains"],
        item["hordes"], item["levels"],
    ] for item in detail_objects]

    by_location = defaultdict(list)
    for item in detail_objects:
        by_location[(item["region"], item["location"])].append(item)

    ranking_objects = []
    for (region, location), items in by_location.items():
        families = {}
        for item in items:
            current = families.get(item["family"])
            if current is None or item["score"] > current["score"]:
                families[item["family"]] = item
        tier_counts = defaultdict(int)
        for item in families.values():
            tier_counts[item["tier"]] += 1
        weighted_score = sum(item["score"] for item in families.values())
        numeric_rates = [item["rate"] for item in items if item["rate"] is not None]
        sorted_items = sorted(items, key=lambda item: (TIER_ORDER[item["tier"]], -item["score"], item["pokemon"]))
        species_text = "；".join(
            f'{item["pokemon"]}({item["tier"]},{item["rate_text"]})' for item in sorted_items
        )
        ranking_objects.append({
            "region": region,
            "location": location,
            "family_count": len(families),
            "species_count": len({item["pokemon_id"] for item in items}),
            "tier_counts": tier_counts,
            "score": weighted_score,
            "night_count": sum(1 for item in items if item["night_only"]),
            "max_rate": max(numeric_rates, default=None),
            "terrains": "、".join(sorted({terrain for item in items for terrain in item["terrains"].split("、")})),
            "species": species_text,
        })

    ranking_objects.sort(key=lambda item: (
        -item["family_count"], -item["score"], -item["species_count"],
        -(item["max_rate"] if item["max_rate"] is not None else -1),
        REGION_ORDER.get(item["region"], 99), item["location"],
    ))
    ranking_rows = []
    for rank, item in enumerate(ranking_objects, 1):
        ranking_rows.append([
            rank, item["region"], item["location"], item["family_count"], item["species_count"],
            item["tier_counts"]["T0"], item["tier_counts"]["T1"], item["tier_counts"]["T2"], item["tier_counts"]["T3"],
            item["score"], item["night_count"], item["max_rate"], item["terrains"], item["species"],
        ])
    return ranking_rows, detail_rows


def setup_sheet(ws, title: str, subtitle: str, headers: list[str], rows: list[list], table_name: str, widths: list[int]):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, title)
    title_cell.font = Font(name="Microsoft YaHei", size=16, bold=True, color=COLORS["white"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    subtitle_cell = ws.cell(2, 1, subtitle)
    subtitle_cell.font = Font(name="Microsoft YaHei", size=10, color=COLORS["text"])
    subtitle_cell.fill = PatternFill("solid", fgColor=COLORS["blue"])
    subtitle_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34

    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(name="Microsoft YaHei", bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["teal"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 32

    for row_index, row in enumerate(rows, header_row + 1):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row_index, col_index, value)
            cell.font = Font(name="Microsoft YaHei", size=9, color=COLORS["text"])
            cell.alignment = Alignment(vertical="top", wrap_text=col_index in {4, 13, 21, 22})
            cell.border = Border(bottom=Side(style="hair", color=COLORS["border"]))

    if rows:
        table = Table(displayName=table_name, ref=f"A{header_row}:{ws.cell(header_row + len(rows), len(headers)).coordinate}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(header_row + len(rows), len(headers)).coordinate}"
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(header_row + len(rows), len(headers)).coordinate}"


def style_selection_sheet(ws, rate_col: int, night_col: int, tier_col: int, rows_count: int):
    if not rows_count:
        return
    start, end = 5, 4 + rows_count
    rate_letter = ws.cell(4, rate_col).column_letter
    ws.conditional_formatting.add(
        f"{rate_letter}{start}:{rate_letter}{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="5B9BD5"),
    )
    for row in range(start, end + 1):
        if ws.cell(row, night_col).value == "是":
            ws.cell(row, night_col).fill = PatternFill("solid", fgColor=COLORS["night"])
            ws.cell(row, night_col).font = Font(name="Microsoft YaHei", size=9, bold=True, color="7030A0")
        tier = ws.cell(row, tier_col).value
        if tier in TIER_COLORS:
            ws.cell(row, tier_col).fill = PatternFill("solid", fgColor=TIER_COLORS[tier])
            ws.cell(row, tier_col).font = Font(name="Microsoft YaHei", size=9, bold=True)
        if ws.cell(row, rate_col).value is not None:
            ws.cell(row, rate_col).number_format = '0.00"%"'


def build_summary_sheet(ws, global_rows: list[list], ranking_rows: list[list], detail_rows: list[list]):
    ws.title = "说明与摘要"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "PokeMMO 野外分布分析"
    ws["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:H2")
    ws["A2"] = f"生成日期：{date.today().isoformat()}  |  数据源：web/search-data.js"
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["blue"])
    ws["A2"].font = Font(name="Microsoft YaHei", size=10)

    ws["A4"] = "统计口径"
    ws["A4"].font = Font(name="Microsoft YaHei", size=13, bold=True, color=COLORS["navy"])
    notes = [
        "全局进化链优选：每条进化链只保留一条最高数值概率记录，并分配到该记录所在地区。",
        "各地区进化链优选：同一进化链可在不同地区各保留一条，便于比较跨地区最佳点。",
        "排序：先按地区，再按出现时间权重；仅夜晚=100（最高），含夜晚但非全天=80，全天=60，其他更低。",
        "选择：数值概率优先；同概率时仅夜晚优先，再优先普通遭遇。香水等无数值记录单独标注。",
        "稀有地点：使用项目现有闪战分级 T0-T3；以不同进化链数量为主、分值合计为次，进化链内不重复计数。",
        "概率是各遭遇方式内部的记录值；普通、3/5群怪和甜甜香气的概率不代表相同触发成本。",
    ]
    for index, note in enumerate(notes, 5):
        ws.merge_cells(start_row=index, start_column=1, end_row=index, end_column=8)
        ws.cell(index, 1, f"• {note}")
        ws.cell(index, 1).font = Font(name="Microsoft YaHei", size=10, color=COLORS["text"])
        ws.cell(index, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[index].height = 24

    summary_start = 13
    ws.cell(summary_start, 1, "地区分配摘要")
    ws.cell(summary_start, 1).font = Font(name="Microsoft YaHei", size=13, bold=True, color=COLORS["navy"])
    headers = ["地区", "全局优选进化链数", "仅夜晚数", "数值概率记录数", "平均最高概率(%)", "最高概率(%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(summary_start + 1, col, header)
        cell.fill = PatternFill("solid", fgColor=COLORS["teal"])
        cell.font = Font(name="Microsoft YaHei", bold=True, color=COLORS["white"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    by_region = defaultdict(list)
    for row in global_rows:
        by_region[row[1]].append(row)
    summary_rows = []
    for region in sorted(by_region, key=lambda name: (REGION_ORDER.get(name, 99), name)):
        rows = by_region[region]
        numeric = [row[6] for row in rows if row[6] is not None]
        summary_rows.append([
            region, len(rows), sum(row[11] == "是" for row in rows), len(numeric),
            sum(numeric) / len(numeric) if numeric else None,
            max(numeric) if numeric else None,
        ])
    for row_index, row in enumerate(summary_rows, summary_start + 2):
        for col_index, value in enumerate(row, 1):
            ws.cell(row_index, col_index, value)
            ws.cell(row_index, col_index).font = Font(name="Microsoft YaHei", size=10)
            ws.cell(row_index, col_index).alignment = Alignment(horizontal="center")
        ws.cell(row_index, 5).number_format = '0.00"%"'
        ws.cell(row_index, 6).number_format = '0.00"%"'

    top_start = 22
    ws.cell(top_start, 1, "稀有地点 TOP 15")
    ws.cell(top_start, 1).font = Font(name="Microsoft YaHei", size=13, bold=True, color=COLORS["navy"])
    top_headers = ["名次", "地区", "地点", "稀有进化链数", "稀有精灵种数", "加权分"]
    for col, header in enumerate(top_headers, 1):
        cell = ws.cell(top_start + 1, col, header)
        cell.fill = PatternFill("solid", fgColor=COLORS["teal"])
        cell.font = Font(name="Microsoft YaHei", bold=True, color=COLORS["white"])
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(ranking_rows[:15], top_start + 2):
        values = [row[0], row[1], row[2], row[3], row[4], row[9]]
        for col_index, value in enumerate(values, 1):
            ws.cell(row_index, col_index, value)
            ws.cell(row_index, col_index).font = Font(name="Microsoft YaHei", size=9)

    chart = BarChart()
    chart.title = "各地区全局优选进化链数"
    chart.y_axis.title = "数量"
    chart.x_axis.title = "地区"
    data_ref = Reference(ws, min_col=2, min_row=summary_start + 1, max_row=summary_start + 1 + len(summary_rows))
    category_ref = Reference(ws, min_col=1, min_row=summary_start + 2, max_row=summary_start + 1 + len(summary_rows))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(category_ref)
    chart.height = 7
    chart.width = 12
    chart.legend = None
    ws.add_chart(chart, "H13")

    for col, width in enumerate([12, 18, 46, 18, 18, 16, 3, 14, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"
    ws["A40"] = f"工作簿包含 {len(global_rows)} 条全局进化链优选、{len(ranking_rows)} 个稀有地点、{len(detail_rows)} 条稀有地点精灵明细。"
    ws["A40"].font = Font(name="Microsoft YaHei", italic=True, color="667085")


def create_workbook():
    data = load_data()
    id_to_family, family_ids, family_names = build_families(data)
    tiers = build_tiers(data, id_to_family, family_ids)
    global_rows, region_rows, special_rows, _ = build_selection_tables(data, id_to_family, family_names, tiers)
    rare_ranking_rows, rare_detail_rows = build_rare_tables(data, id_to_family, family_names, tiers)

    selection_headers = [
        "地区排名", "地区", "进化链编号", "进化链", "选中精灵", "图鉴编号",
        "最高概率(%)", "概率显示", "最高概率时段", "出现时间", "时间权重", "仅夜晚",
        "地点", "地形/方式", "等级", "季节", "遭遇类型", "闪战分级", "闪战分值",
        "覆盖地区数", "其他可选地区", "选择依据",
    ]
    selection_widths = [10, 10, 12, 38, 14, 10, 14, 12, 16, 14, 10, 10, 40, 16, 12, 10, 12, 12, 12, 12, 28, 38]

    wb = Workbook()
    build_summary_sheet(wb.active, global_rows, rare_ranking_rows, rare_detail_rows)

    ws_global = wb.create_sheet("全局进化链优选")
    setup_sheet(
        ws_global,
        "全局进化链优选",
        "每条进化链全局仅保留一个最佳记录，并归入其最佳地区；地区内仅夜晚出现的记录排在最前。",
        selection_headers,
        global_rows,
        "GlobalEvolutionBest",
        selection_widths,
    )
    style_selection_sheet(ws_global, 7, 12, 18, len(global_rows))

    ws_region = wb.create_sheet("各地区进化链优选")
    region_headers = selection_headers + ["是否全局分配地区"]
    setup_sheet(
        ws_region,
        "各地区进化链优选",
        "每条进化链在每个有分布的地区各保留一个最佳记录，用于横向比较地区 A/B 的最高概率。",
        region_headers,
        region_rows,
        "RegionalEvolutionBest",
        selection_widths + [18],
    )
    style_selection_sheet(ws_region, 7, 12, 18, len(region_rows))

    rare_headers = [
        "名次", "地区", "地点", "稀有进化链数", "稀有精灵种数",
        "T0链数", "T1链数", "T2链数", "T3链数", "稀有度加权分",
        "仅夜晚精灵数", "最高概率(%)", "主要地形/方式", "稀有精灵清单（等级,最高概率）",
    ]
    ws_rare = wb.create_sheet("稀有地点排行")
    setup_sheet(
        ws_rare,
        "稀有地点排行",
        "稀有口径为闪战分级 T0-T3；优先按不同进化链数量排序，再按分值合计与精灵种数排序。",
        rare_headers,
        rare_ranking_rows,
        "RareLocationRanking",
        [9, 10, 44, 16, 16, 10, 10, 10, 10, 16, 16, 14, 28, 90],
    )
    if rare_ranking_rows:
        end = 4 + len(rare_ranking_rows)
        ws_rare.conditional_formatting.add(
            f"J5:J{end}",
            DataBarRule(start_type="min", end_type="max", color="F4B183"),
        )
        for row in range(5, end + 1):
            if ws_rare.cell(row, 12).value is not None:
                ws_rare.cell(row, 12).number_format = '0.00"%"'

    rare_detail_headers = [
        "地区", "地点", "精灵", "图鉴编号", "进化链编号", "进化链", "闪战分级", "分值",
        "最高概率(%)", "概率显示", "最佳时段", "出现时间汇总", "仅夜晚", "季节汇总",
        "地形/方式汇总", "遭遇类型汇总", "等级样例",
    ]
    ws_detail = wb.create_sheet("稀有地点明细")
    setup_sheet(
        ws_detail,
        "稀有地点明细",
        "按地区、地点和精灵聚合 T0-T3 记录；同一地点的季节、时间、方式与等级已合并。",
        rare_detail_headers,
        rare_detail_rows,
        "RareLocationDetail",
        [10, 42, 16, 10, 12, 38, 12, 10, 14, 12, 14, 18, 10, 22, 28, 20, 28],
    )
    style_selection_sheet(ws_detail, 9, 13, 7, len(rare_detail_rows))

    ws_special = wb.create_sheet("特殊遇见")
    setup_sheet(
        ws_special,
        "无数值概率的特殊遇见",
        "这些进化链在对应地区没有可比较的百分比记录，通常为“香水”等特殊遇见，因此不参与数值概率竞争。",
        selection_headers,
        special_rows,
        "SpecialEncounterBest",
        selection_widths,
    )
    style_selection_sheet(ws_special, 7, 12, 18, len(special_rows))

    wb.properties.title = "PokeMMO 进化链与稀有地点分析"
    wb.properties.subject = "野外出现概率、时间权重、地区分配与稀有地点统计"
    wb.properties.creator = "Codex"
    wb.save(OUTPUT)

    # Reopen once to ensure the generated OOXML package is readable.
    checked = load_workbook(OUTPUT, read_only=True, data_only=False)
    sheet_names = checked.sheetnames
    checked.close()
    print(json.dumps({
        "output": str(OUTPUT),
        "records": len(data["r"]),
        "global_rows": len(global_rows),
        "regional_rows": len(region_rows),
        "rare_locations": len(rare_ranking_rows),
        "rare_details": len(rare_detail_rows),
        "special_rows": len(special_rows),
        "sheets": sheet_names,
    }, ensure_ascii=False))


if __name__ == "__main__":
    create_workbook()
